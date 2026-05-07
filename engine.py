import requests, pandas as pd, re, smtplib, json, os
from datetime import date, datetime, timezone, timedelta
from openpyxl.styles import PatternFill, Font, Alignment
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

USERNAME = "terrycook6"
PASSWORD = "#Dolour01"
APP_KEY = "4CEtH9teFpZzYtoG"
RACING_USER = "EPkM7mogtxBREaZloFZuExEM"
RACING_PASS = "1Owdip14JEoyxl0aT7Z7tMYp"
BASE = "https://api.theracingapi.com/v1/australia"
GMAIL = "terrycook230759@gmail.com"
APP_PASSWORD = "gvmpdjrszwmunfbv"
TO_EMAIL = "terrycook6@msn.com"

# Minutes before race to stop including it
CUTOFF_MINUTES = 5

# Only send emails after this hour (8:30am = 8.5)
EMAIL_START_HOUR = 8.5

# Only resend email if edge changes by more than this amount
EMAIL_EDGE_CHANGE_THRESHOLD = 1.5

BLOCKED_TRACKS = ["lark hill","tabcorp park","melton","penrith park","harold park"]
BLOCKED_CLASSES = ["TRL","TRIAL","JUMP OUT","JUMPOUT","HDL-TRL","STP TRL","OPEN-BT","2YO TRL","3YO TRL","-BT","BT","OPEN-TRL","MDN-TRL","CL1-TRL","CL2-TRL","CL3-TRL","CL4-TRL","CL5-TRL","2Y-TRL","3Y-TRL","MDN-CL","OPEN"]

TRACK_LOOKUP = {
    "picklebet park wodonga": "wodonga",
    "bet365 terang": "terang",
    "sportsbet mareeba": "mareeba",
    "tab toowoomba": "toowoomba",
    "tab rockhampton": "rockhampton",
    "tab townsville": "townsville",
    "tab gold coast": "gold coast",
    "tab canberra": "canberra",
    "southside cranbourne": "cranbourne",
}

def clean(name):
    name = name.lower().replace(chr(39),"").replace(chr(8217),"").replace(".","").strip()
    name = re.sub(r"^[0-9]+ ", "", name)
    return name

def get_bf_track_name(course):
    course_lower = course.lower()
    if course_lower in TRACK_LOOKUP:
        return TRACK_LOOKUP[course_lower]
    for prefix in ["picklebet park ","bet365 ","sportsbet ","tab ","tabcorp "]:
        course_lower = course_lower.replace(prefix, "")
    return course_lower.strip()

def parse_off_time(off_time_str):
    try:
        if off_time_str:
            dt_str = off_time_str[:19]
            return datetime.strptime(dt_str, "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)
    except:
        pass
    return None

def is_race_upcoming(off_time_str):
    off_time = parse_off_time(off_time_str)
    if off_time is None:
        return True
    now_utc = datetime.now(timezone.utc)
    minutes_until = (off_time - now_utc).total_seconds() / 60
    return minutes_until > CUTOFF_MINUTES

def format_race_time(off_time_str):
    off_time = parse_off_time(off_time_str)
    if off_time is None:
        return ""
    aest = off_time + timedelta(hours=10)
    return aest.strftime("%I:%M %p")

def is_email_time():
    """Only send emails after EMAIL_START_HOUR"""
    now_aest = datetime.now(timezone.utc) + timedelta(hours=10)
    current_hour = now_aest.hour + now_aest.minute / 60
    return current_hour >= EMAIL_START_HOUR

def load_emailed(emailed_file):
    """Load previously emailed bets for today"""
    try:
        if os.path.exists(emailed_file):
            data = json.load(open(emailed_file))
            # Reset if it's a new day
            if data.get("date") != str(date.today()):
                return {"date": str(date.today()), "bets": {}}
            return data
    except:
        pass
    return {"date": str(date.today()), "bets": {}}

def save_emailed(emailed_file, data):
    """Save emailed bets record"""
    try:
        json.dump(data, open(emailed_file, "w"))
    except:
        pass

def filter_new_value_bets(value_rows, emailed_data):
    """Only return bets that are new or have significantly changed edge"""
    new_bets = []
    for r in value_rows:
        # Create unique key for this bet
        key = f"{r.get('Track','')}_{r.get('Race/No','').replace(chr(10),'_')}_{r.get('Horse','')}"
        edge_bf = float(r.get('Edge vs BF') or 0)
        edge_tab = float(r.get('Edge vs TAB') or 0)
        current_edge = max(edge_bf, edge_tab)

        prev = emailed_data["bets"].get(key)
        if prev is None:
            # New bet - include it
            new_bets.append(r)
            emailed_data["bets"][key] = current_edge
        else:
            # Already emailed - only resend if edge changed significantly
            if abs(current_edge - prev) >= EMAIL_EDGE_CHANGE_THRESHOLD:
                new_bets.append(r)
                emailed_data["bets"][key] = current_edge
    return new_bets

STEAM_DROP_THRESHOLD = 0.20   # 20% price drop triggers alert
STEAM_MAX_PRICE = 15.0        # Ignore horses over this price

def load_odds_history():
    """Load previous Betfair odds from file"""
    try:
        if os.path.exists('odds_history.json'):
            data = json.load(open('odds_history.json'))
            if data.get('date') != str(date.today()):
                return {'date': str(date.today()), 'odds': {}}
            return data
    except:
        pass
    return {'date': str(date.today()), 'odds': {}}

def save_odds_history(data):
    try:
        json.dump(data, open('odds_history.json', 'w'))
    except:
        pass

def check_steam(track, race_no, horse, current_price, odds_history):
    """Check if horse has steamed since last run"""
    key = f"{track}_{race_no}_{horse}"
    prev_price = odds_history['odds'].get(key)
    odds_history['odds'][key] = current_price
    if prev_price and current_price and current_price <= STEAM_MAX_PRICE:
        drop = (prev_price - current_price) / prev_price
        if drop >= STEAM_DROP_THRESHOLD:
            return True, prev_price, current_price, round(drop * 100, 1)
    return False, None, None, None

def send_email(value_rows):
    if not value_rows:
        print('No new value bets to email')
        return
    parts = ['VALUE BETS - UPCOMING RACES', '']
    for r in value_rows:
        race_time = r.get('Race Time', '')
        parts.append(str(r.get('Track','')) + ' ' + str(r.get('Race/No','').replace(chr(10),' ')) + ' - ' + str(r.get('Horse','')))
        if race_time:
            parts.append('  Race Time: ' + race_time)
        parts.append('  Win%: ' + str(r.get('Win%','')) + '%  Rating: ' + str(r.get('Rating','')))
        parts.append('  True Odds: $' + str(r.get('True Odds','')))
        parts.append('  TAB: $' + str(r.get('TAB Odds','')) + '  Betfair: $' + str(r.get('Betfair Odds','')))
        parts.append('  Edge TAB: ' + str(r.get('Edge vs TAB','')) + '%  Edge BF: ' + str(r.get('Edge vs BF','')) + '%')
        parts.append('  ' + str(r.get('Comment','')))
        parts.append('')
    body = chr(10).join(parts)
    msg = MIMEMultipart()
    msg['From'] = GMAIL
    msg['To'] = TO_EMAIL
    msg['Subject'] = 'Racing Value Bets ' + str(date.today())
    msg.attach(MIMEText(body, 'plain'))
    try:
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(GMAIL, APP_PASSWORD)
        server.sendmail(GMAIL, TO_EMAIL, msg.as_string())
        server.quit()
        print('Email sent with ' + str(len(value_rows)) + ' value bets')
    except Exception as e:
        print('Email failed: ' + str(e))

def send_steam_email(steam_rows):
    if not steam_rows:
        return
    parts = ['STEAM ALERT — LARGE MONEY DETECTED', '']
    for r in steam_rows:
        parts.append(str(r.get('Track','')) + ' ' + str(r.get('Race/No','').replace(chr(10),' ')) + ' - ' + str(r.get('Horse','')))
        parts.append(f"  Race Time: {r.get('Race Time','')}")
        parts.append(f"  Price moved: ${r.get('prev_price','')} → ${r.get('current_price','')} ({r.get('drop_pct','')}% drop)")
        parts.append(f"  Win%: {r.get('Win%','')}%  Edge vs BF: {r.get('Edge vs BF','')}%")
        parts.append(f"  {r.get('Comment','')}")
        parts.append('')
    body = chr(10).join(parts)
    msg = MIMEMultipart()
    msg['From'] = GMAIL
    msg['To'] = TO_EMAIL
    msg['Subject'] = 'STEAM ALERT ' + str(date.today())
    msg.attach(MIMEText(body, 'plain'))
    try:
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(GMAIL, APP_PASSWORD)
        server.sendmail(GMAIL, TO_EMAIL, msg.as_string())
        server.quit()
        print(f'Steam alert sent for {len(steam_rows)} horses')
    except Exception as e:
        print(f'Steam email failed: {e}')

def betfair_login():
    r = requests.post("https://identitysso.betfair.com/api/login",
        data={"username": USERNAME, "password": PASSWORD},
        headers={"X-Application": APP_KEY, "Accept": "application/json"})
    return r.json()["token"]

def get_betfair_odds(token, track, race_number):
    try:
        resp = requests.post("https://api.betfair.com/exchange/betting/json-rpc/v1",
            json=[{"jsonrpc":"2.0","method":"SportsAPING/v1.0/listMarketCatalogue","params":{"filter":{"eventTypeIds":["7"],"marketCountries":["AU"],"marketTypeCodes":["WIN"],"textQuery":track},"maxResults":10,"marketProjection":["RUNNER_DESCRIPTION","EVENT"]},"id":1}],
            headers={"X-Application": APP_KEY, "X-Authentication": token, "Content-Type": "application/json"})
        markets = resp.json()[0]["result"]
        market = None
        for m in markets:
            if "R" + str(race_number) in m["marketName"]:
                market = m
                break
        if not market:
            return {}
        id_to_name = {r["selectionId"]: r["runnerName"] for r in market["runners"]}
        odds_resp = requests.post("https://api.betfair.com/exchange/betting/json-rpc/v1",
            json=[{"jsonrpc":"2.0","method":"SportsAPING/v1.0/listMarketBook","params":{"marketIds":[market["marketId"]],"priceProjection":{"priceData":["EX_BEST_OFFERS"]}},"id":1}],
            headers={"X-Application": APP_KEY, "X-Authentication": token, "Content-Type": "application/json"})
        runners = odds_resp.json()[0]["result"][0]["runners"]
        result = {}
        for r in runners:
            best = r.get("ex", {}).get("availableToBack", [])
            price = best[0]["price"] if best else None
            name = id_to_name.get(r["selectionId"], "")
            if name and price:
                result[clean(name)] = price
        return result
    except:
        return {}

def get_meets():
    r = requests.get(f"{BASE}/meets", params={"date": str(date.today())}, auth=(RACING_USER, RACING_PASS))
    return r.json().get("meets", [])

def get_race(meet_id, race_number):
    r = requests.get(f"{BASE}/meets/{meet_id}/races/{race_number}", auth=(RACING_USER, RACING_PASS))
    return r.json()

def get_odds(runner):
    for o in runner.get("odds", []):
        try:
            p = float(o.get("win_odds", 0))
            if p > 1:
                return p
        except:
            pass
    return None

def rate_runner(runner, going):
    odds = get_odds(runner)
    base = round(100 / odds, 2) if odds and odds > 1 else 10.0
    stats = runner.get("stats", {})
    d = stats.get("distance_stats", {})
    if int(d.get("total", 0)) > 0:
        base += int(d.get("first", 0)) / int(d.get("total", 1)) * 5
    c = stats.get("course_stats", {})
    if int(c.get("total", 0)) > 0:
        base += int(c.get("first", 0)) / int(c.get("total", 1)) * 4
    going_lower = going.lower()
    if "soft" in going_lower or "heavy" in going_lower:
        g = stats.get("ground_heavy_stats") or {} if "heavy" in going_lower else stats.get("ground_soft_stats") or {}
        g_runs = int(g.get("total", 0))
        g_wins = int(g.get("first", 0))
        if g_runs >= 2:
            g_rate = g_wins / g_runs
            base += (g_rate - 0.15) * 10
    try:
        claim = int(runner.get("jockey_claim", 0))
        base -= claim * 0.5
    except:
        pass
    try:
        last_raced = runner.get("last_raced")
        if last_raced:
            days = (datetime.today() - datetime.strptime(last_raced[:10], "%Y-%m-%d")).days
            if days > 90: base -= 2
            elif days > 60: base -= 1
            elif 14 <= days <= 28: base += 1
    except:
        pass
    return round(max(base, 1), 2)

def main():
    print("Logging into Betfair...")
    token = betfair_login()
    print("Done")
    meets = get_meets()
    print(f"Found {len(meets)} meets today")
    all_rows = []
    for meet in meets:
        course = meet.get("course", "")
        races = meet.get("races", [])
        if any(course.lower() == b or course.lower().startswith(b + " ") for b in BLOCKED_TRACKS):
            continue
        real_races = [r for r in races if not any(x in r.get("class","").upper() for x in BLOCKED_CLASSES) and not r.get("class","").strip().endswith("-")]
        if not real_races:
            continue
        upcoming_races = [r for r in real_races if is_race_upcoming(r.get("off_time",""))]
        if not upcoming_races:
            continue
        print(f"  {course} - {len(upcoming_races)} upcoming races")
        bf_name = get_bf_track_name(course)
        for race_info in upcoming_races:
            try:
                race_number = race_info["race_number"]
                off_time = race_info.get("off_time","")
                race_time = format_race_time(off_time)
                off_time_dt = parse_off_time(off_time)
                race = get_race(meet["meet_id"], race_number)
                going = race.get("going") or "Good"
                runners = [r for r in race.get("runners", []) if not r.get("scratched", False)]
                if not runners:
                    continue
                bf_odds = get_betfair_odds(token, bf_name, race_number)
                scores = [(r, rate_runner(r, going)) for r in runners]
                total = sum(s for _, s in scores)
                for r, score in sorted(scores, key=lambda x: int(x[0].get("number","99") if str(x[0].get("number","99")).isdigit() else 99)):
                    prob = round(score / total * 100, 1)
                    true_odd = round(100 / prob, 2)
                    mkt = get_odds(r)
                    mkt_prob = round(100 / mkt, 1) if mkt else None
                    edge_tab = round(prob - mkt_prob, 1) if mkt_prob else None
                    horse_name = r.get("horse", "")
                    draw = r.get("draw", "")
                    horse_num = r.get("number", "")
                    horse_with_draw = f"{horse_name} ({draw})" if draw else horse_name
                    jockey = r.get("jockey", "")
                    trainer = r.get("trainer", "")
                    jockey_trainer = (jockey or "") + chr(10) + (trainer or "")
                    race_horse_no = "R" + str(race_number) + chr(10) + "H" + str(horse_num)
                    bf_price = bf_odds.get(clean(horse_name))
                    bf_prob = round(100 / bf_price, 1) if bf_price else None
                    edge_bf = round(prob - bf_prob, 1) if bf_price else None
                    rating = min(99, max(1, round(prob * 3 + max(0,(edge_tab or 0)) * 2 + max(0,(edge_bf or 0)) * 2)))
                    all_rows.append({
                        "Race Time": race_time,
                        "off_time_dt": off_time_dt,
                        "Track": course,
                        "Race/No": race_horse_no,
                        "Distance": race_info.get("distance",""),
                        "Going": going,
                        "Horse": horse_with_draw,
                        "Jockey/Trainer": jockey_trainer,
                        "Comment": r.get("comment",""),
                        "Win%": prob,
                        "Rating": rating,
                        "True Odds": true_odd,
                        "TAB Odds": mkt,
                        "Betfair Odds": bf_price,
                        "Edge vs TAB": edge_tab,
                        "Edge vs BF": edge_bf,
                        "Value TAB": "YES" if edge_tab and edge_tab > 3.0 else "",
                        "Value BF": "YES" if edge_bf and edge_bf > 3.0 else ""
                    })
            except Exception as e:
                print(f"    Error: {e}")
                continue

    if all_rows:
        # Load odds history for steam detection
        odds_history = load_odds_history()
        steam_rows = []

        # Check for steam moves
        for r in all_rows:
            bf_price = r.get('Betfair Odds')
            if bf_price:
                is_steam, prev_price, current_price, drop_pct = check_steam(
                    r.get('Track',''), r.get('Race/No',''), r.get('Horse',''),
                    bf_price, odds_history
                )
                if is_steam:
                    steam_rows.append({**r, 'prev_price': prev_price, 'current_price': current_price, 'drop_pct': drop_pct})

        save_odds_history(odds_history)
        # Sort by race time - next to go first
        all_rows.sort(key=lambda x: x.get("off_time_dt") or datetime.max.replace(tzinfo=timezone.utc))

        # Remove sort key before saving
        for row in all_rows:
            row.pop("off_time_dt", None)

        df = pd.DataFrame(all_rows)
        fname = f"reports/racing_{date.today()}.xlsx"
        writer = pd.ExcelWriter(fname, engine="openpyxl")
        df.to_excel(writer, index=False, sheet_name="Races")
        ws = writer.sheets["Races"]
        ws.auto_filter.ref = "A1:B1"
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            if cell.value == "Value BF":
                cell.value = "Value" + chr(10) + "BF"
            if cell.value == "Value TAB":
                cell.value = "Value" + chr(10) + "TAB"
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_widths = {"A":9,"B":10,"C":11,"D":10,"E":7,"F":25,"G":28,"H":50,"I":5,"J":6,"K":8,"L":8,"M":8,"N":8,"O":8,"P":8,"Q":8}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        for cell in ws[1]:
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.alignment = wrap_center
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = wrap_center
        green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        headers = [cell.value for cell in ws[1]]
        bf_col  = headers.index("Edge vs BF")  + 1 if "Edge vs BF"  in headers else None
        tab_col = headers.index("Edge vs TAB") + 1 if "Edge vs TAB" in headers else None
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            bf_val  = row[bf_col-1].value  if bf_col  else None
            tab_val = row[tab_col-1].value if tab_col else None
            row_fill = None
            try:
                if bf_val  and float(bf_val)  > 3.0: row_fill = green
                elif tab_val and float(tab_val) > 3.0: row_fill = green
                elif bf_val  and float(bf_val)  > 0: row_fill = yellow
                elif tab_val and float(tab_val) > 0: row_fill = yellow
            except:
                pass
            if row_fill:
                for cell in row:
                    cell.fill = row_fill
        race_colours = ["FFF2CC","DDEBF7","E2EFDA","FCE4D6","EAD1DC","D9E1F2","EDEDED","FCE4D6","DDEBF7","E2EFDA","FFF2CC","D9E1F2"]
        b_col = headers.index("Race/No") + 1 if "Race/No" in headers else None
        if b_col:
            current_race = None
            colour_index = -1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[b_col-1]
                val = str(cell.value or "")
                race_id = val.split(chr(10))[0] if chr(10) in val else val
                if race_id != current_race:
                    current_race = race_id
                    colour_index = (colour_index + 1) % len(race_colours)
                col_fill = PatternFill(start_color=race_colours[colour_index], end_color=race_colours[colour_index], fill_type="solid")
                cell.fill = col_fill

        # Handle emails - only after 8:30am and only new/changed value bets
        value_rows = [r for r in all_rows if r.get('Value TAB') == 'YES' or float(r.get('Edge vs BF') or 0) > 3.0]

        if is_email_time():
            emailed_file = 'emailed.json'
            emailed_data = load_emailed(emailed_file)
            new_value_rows = filter_new_value_bets(value_rows, emailed_data)
            if new_value_rows:
                send_email(new_value_rows)
                save_emailed(emailed_file, emailed_data)
            else:
                print('No new value bets to email')
        else:
            print('Before 8:30am - email alerts not active yet')

        writer.close()
        print(f"Done - saved to {fname}")
        print(f"Total upcoming runners: {len(all_rows)}")
    else:
        print("No upcoming races found")

if __name__ == "__main__":
    main()
